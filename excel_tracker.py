import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime
from typing import Dict, List
import os

class ExcelTracker:
    """Update your job application Excel tracker"""
    
    COLUMNS = [
        "Date Found", "Company", "Role", "Location", 
        "Fitment Score", "Recommendation", "JD URL",
        "Salary", "Status", "Referral Available", 
        "Connection Name", "Applied Date", "Notes"
    ]
    
    COLORS = {
        "APPLY": "C6EFCE",    # green
        "CONSIDER": "FFEB9C", # yellow  
        "SKIP": "FFC7CE"      # red
    }
    
    def __init__(self, file_path: str = "job_tracker.xlsx"):
        self.file_path = file_path
        self._ensure_file_exists()
    
    def _ensure_file_exists(self):
        if not os.path.exists(self.file_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Applications"
            
            # Header row
            for col, header in enumerate(self.COLUMNS, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            
            wb.save(self.file_path)
    
    def add_jobs(self, scored_jobs: List[Dict], referral_data: Dict = None) -> int:
        """
        Add shortlisted jobs to tracker.
        Only adds jobs not already in the sheet (dedupes by company+role).
        Returns count of new rows added.
        """
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb["Applications"]
        
        # Build set of existing entries to avoid duplicates
        existing = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] and row[2]:  # company + role
                existing.add((str(row[1]).lower(), str(row[2]).lower()))
        
        added = 0
        for job in scored_jobs:
            key = (job.get('company','').lower(), job.get('job_title','').lower())
            if key in existing:
                continue
            
            referral = referral_data.get(job.get('company'), {}) if referral_data else {}
            
            row_data = [
                datetime.now().strftime("%Y-%m-%d"),
                job.get('company'),
                job.get('job_title'),
                job.get('location'),
                job.get('overall_score'),
                job.get('recommendation'),
                job.get('job_url'),
                job.get('salary'),
                'Not Applied',
                'Yes' if referral.get('has_connection') else 'No',
                referral.get('connection_name', ''),
                '',  # applied date — you fill this
                ''   # notes — you fill this
            ]
            
            row_num = ws.max_row + 1
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col, value=value)
            
            # Color code by recommendation
            rec = job.get('recommendation', 'SKIP')
            fill_color = self.COLORS.get(rec, "FFFFFF")
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            for col in range(1, len(self.COLUMNS) + 1):
                ws.cell(row=row_num, column=col).fill = fill
            
            existing.add(key)
            added += 1
        
        wb.save(self.file_path)
        return added
